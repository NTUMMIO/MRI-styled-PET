import numpy as np
import os
from PIL import Image
from skimage.metrics import structural_similarity as ssim
import cv2
import math
from scipy import stats
import csv
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
import argparse

import pandas as pd
import statsmodels.api as sm
from statsmodels.formula.api import ols 
from statsmodels.stats.multicomp import pairwise_tukeyhsd

def t_test(group1, group2):
    group1_mean=np.mean(group1)
    group2_mean=np.mean(group2)
    group1_std=np.std(group1)
    group2_std=np.std(group2)
    
    group1_samples=len(group1)
    group2_samples=len(group2)
    print("group1",group1_mean,group1_std,group1_samples)
    print("group2",group2_mean,group2_std,group2_samples)
    modified_std1 = np.sqrt(np.float32(group1_samples)/np.float32(group1_samples)) * group1_std
    modified_std2 = np.sqrt(np.float32(group2_samples)/np.float32(group2_samples)) * group2_std

    (statistic, pvalue) = stats.ttest_ind_from_stats(mean1=group1_mean, std1= modified_std1, nobs1=group1_samples, mean2=group2_mean, std2= modified_std2, nobs2=group2_samples)
    return statistic, pvalue

def subject_mean(mat):
    subject_list=[]
    print(len(mat))
    for i in range(10):
        if i==7:
            mean=np.mean(mat[248*i:248*i+247-1],axis=0)
            #print(248*i,248*i+247-1, mean)
        if i<7:
            mean=np.mean(mat[248*i:248*i+248-1],axis=0)
            #print(248*i,248*i+248-1, mean)
        if i>7:
            mean=np.mean(mat[248*i-1:248*i+247-1],axis=0)
            #print(248*i-1,248*i+247-1, mean)
        subject_list.append(mean)    
    subject_list=np.array(subject_list)
    print(subject_list.shape)
    return subject_list

if __name__=='__main__':
    
    groups=['AD', 'MCI', 'CN']
            
    # metrics_p1=['SSIM_PET','SSIM_PVC','PSNR_PET','PSNR_PVC']
    metrics_p1=['SSIM_PET','SSIM_PVC','SSIM_MRI','PSNR_PET','PSNR_PVC','PSNR_MRI']


    condition_folders=[
        "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/Baseline/fold0/2024-9-10-113411/AD.csv",

        # "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigA/fold0/2024-9-12-123532/AD.csv",
        # "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigA/fold0/2024-9-15-213741/AD.csv",

        # "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigB_ratio1/fold0/2024-9-12-12336/AD.csv",
        # "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigB_ratio1/fold0/2024-9-15-213720/AD.csv"

        # "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigC_ratio1/fold0/2024-9-12-164835/AD.csv",
        # "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigC_ratio1/fold0/2024-9-15-213750/AD.csv",

        # "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigD_ratio1/fold0/2024-9-12-165847/AD.csv",
        # "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigA/fold0/2024-9-16-191915_train7_inference7/AD.csv",
        # "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigA/fold0/2024-9-16-191915_train7_inference3/AD.csv",
        # "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigA/fold0/2024-9-18-95846_train3_inference7/AD.csv",
        # "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigA/fold0/2024-9-18-95846_train3_inference3/AD.csv",

        "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/Baseline/fold0/Baseline/AD.csv", 
        "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigB_ratio05/fold0/2024-9-16-191956_train7_inference7/AD.csv",
        "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigB_ratio05/fold0/2024-9-16-191956_train7_inference3/AD.csv",
        "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigB_ratio05_anatomical3/fold0/2024-9-18-13836_train3_inference7/AD.csv",
        "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigB_ratio05_anatomical3/fold0/2024-9-18-13836_train3_inference3/AD.csv",

        # "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/Baseline/fold0/Baseline/AD.csv", 
        # "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigA/fold0/ConfigA_real/AD.csv",
        # # "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigA/fold0/ConfigA_pseudo/AD.csv",
        # "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigB/fold0/ConfigB_real/AD.csv",
        # # "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigB/fold0/ConfigB_pseudo/AD.csv",
        # "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigC/fold0/ConfigC_real/AD.csv",
        # "/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/raw_quantitative_metrics/ConfigD/fold0/ConfigD_real/AD.csv"
        
    ]


    
    cross_group_control = {}
    cross_group_experiment = {}
    

    # Iterate through each element in metrics_p1
    for metric in metrics_p1:
        # Assign the desired value to each key
        cross_group_control[metric] = {'AD':[], 'MCI':[], 'CN':[]}
        cross_group_experiment[metric] = {'AD':[], 'MCI':[], 'CN':[]}
        

    for group in groups:

        wb = openpyxl.Workbook()
        
        wb_brief = openpyxl.Workbook()
        ws_brief=wb_brief.create_sheet()


        for filename in condition_folders:
            
            print("============================================")

            df1=pd.read_csv(filename.replace('AD', group)).drop(['Unnamed: 0'],axis=1)#, names=['Subject', 'SSIM_PET','SSIM_PVC','PSNR_PET','PSNR_PVC'])

            print(filename.replace('AD', group))    
            print(df1)

            df = df1

        #=========================================================================    
        #specify calculate metrics on how many fold
            #df = df[:20]
        #=========================================================================  

            for m in range(len(metrics_p1)):

                metrics=metrics_p1[m]
                #print(metrics)       
                try:
                    ws = wb[metrics]
                except KeyError:
                    ws=wb.create_sheet(title=metrics)

                this_col=ws.max_column
                
                ws.cell(row=1, column=this_col+1).value = filename.split('/')[-4]+"_"+filename.split('/')[-2]#[:-15]
                # ws.cell(row=2, column=this_col+1).value = filename.split('/')[-3]+'_'+filename.split('/')[-2]
                
                for i in range(df.shape[0]):
                    ws.cell(row=i+2, column=1).value = df.iloc[i,0]
                    ws.cell(row=i+2, column=this_col+1).value = df.iloc[i,m+1]
        
        
                
        metrics_all=metrics_p1#+metrics_p2
        
        for m in range(len(metrics_all)):
            metrics=metrics_all[m]
            ws = wb[metrics]
            print("===========================================")
            print(metrics)
            
            max_c=ws.max_column
            # print(max_c)
            max_r=ws.max_row
            ws.cell(row=max_r+1, column=1).value = "mean"
            ws.cell(row=max_r+2, column=1).value = "p-value(with control)"
            
            
            ws_brief.cell(row=3*m+2, column=1).value = metrics
            ws_brief.cell(row=1, column=2).value = ws.cell(row=1, column=2).value
            
            for c in range(3,max_c+1):
                # print(c, ws.cell(row=1, column=c).value,ws.cell(row=2, column=c).value)
                control_list=[]
                experiment_list=[]
                for j in range(3, max_r+1):
                    control_list.append(ws.cell(row=j, column=c-1).value)
                    experiment_list.append(ws.cell(row=j, column=c).value)
                    
                    cross_group_control[metrics_all[m]][group]=control_list
                    cross_group_experiment[metrics_all[m]][group]=experiment_list
                    
                print(len(control_list), len(experiment_list))
                print(control_list)
                print(experiment_list)
                # statistic, pvalue=stats.ttest_rel(control_list, experiment_list)
                
                pvalue=stats.wilcoxon(control_list, experiment_list).pvalue
                # print(i,statistic, pvalue)  
                
                mean_control=sum(control_list)/len(control_list)
                mean_experiment=sum(experiment_list)/len(experiment_list)
                
                ws.cell(row=max_r+1, column=c-1).value = mean_control
                ws.cell(row=max_r+1, column=c).value = mean_experiment
                if mean_control<mean_experiment:
                    ws.cell(row=max_r+1, column=c).font=Font(color = "FF0000")
                else:
                    ws.cell(row=max_r+1, column=c).font=Font(color = "0000FF")
                
                
                ws.cell(row=max_r+2, column=c).value = pvalue
                if pvalue<0.05:
                    ws.cell(max_r+3, column=c).value = "*"
                if pvalue<0.01:
                    ws.cell(max_r+3, column=c).value = "**"
                if pvalue<0.001:
                    ws.cell(max_r+3, column=c).value = "***"
                    
                
                ws_brief.cell(row=3*m+3, column=1).value = "p-value"
                ws_brief.cell(row=3*m+2, column=c).value = mean_experiment
                ws_brief.cell(row=3*m+3, column=c).value = pvalue
                ws_brief.cell(row=1, column=c).value = ws.cell(row=1, column=c).value

                if mean_control<mean_experiment:
                    ws_brief.cell(row=3*m+2, column=c).font=Font(color = "FF0000")
                else:
                    ws_brief.cell(row=3*m+2, column=c).font=Font(color = "0000FF")
                
                ws_brief.cell(row=3*m+2, column=c-1).value = mean_control
                if pvalue<0.05:
                    ws_brief.cell(3*m+4, column=c).value = "*"
                if pvalue<0.01:
                    ws_brief.cell(3*m+4, column=c).value = "**"
                if pvalue<0.001:
                    ws_brief.cell(3*m+4, column=c).value = "***"

        if not os.path.exists("/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/metrics_statistics"):
            os.makedirs("/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/metrics_statistics")                        
        wb.save("/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/metrics_statistics/{}.xlsx".format(group))
        wb_brief.save("/home/linyunong/project/MRI-styled-PET-Dual-modality-Fusion-for-PET-Image-Enhancement/analysis/results/metrics_statistics/{}_brief.xlsx".format(group))
        
        


    for metric in metrics_p1:
        if metric.find('MI')==-1:# and metric.find('MRI')==-1:
            print("=================================================")
            print(metric)
            
            
            print(stats.shapiro(cross_group_control[metric]['AD']), stats.shapiro(cross_group_control[metric]['MCI']), stats.shapiro(cross_group_control[metric]['CN']))
            statistic, p_value=stats.levene(cross_group_control[metric]['AD'], cross_group_control[metric]['MCI'], cross_group_control[metric]['CN'])
            print("control", statistic, p_value)
            groups = np.array(['AD'] * len(cross_group_control[metric]['AD']) + ['MCI'] * len(cross_group_control[metric]['MCI']) + ['CN'] * len(cross_group_control[metric]['CN']))
            df = pd.DataFrame({'Data': np.concatenate([cross_group_control[metric]['AD'], cross_group_control[metric]['MCI'], cross_group_control[metric]['CN']]), 'Group': groups})
            model = ols('Data ~ C(Group)', data=df).fit()
            anova_table = sm.stats.anova_lm(model, typ=2)
            print(anova_table)
            if anova_table['PR(>F)']['C(Group)'] < 0.05:
                print("There is a significant difference in means among the groups.")
            else:
                print("There is no significant difference in means among the groups.")
            posthoc_results = pairwise_tukeyhsd(endog=df['Data'], groups=df['Group'], alpha=0.05)
            print(posthoc_results)
            
            
            print(stats.shapiro(cross_group_experiment[metric]['AD']), stats.shapiro(cross_group_experiment[metric]['MCI']), stats.shapiro(cross_group_experiment[metric]['CN']))
            statistic, p_value=stats.levene(cross_group_experiment[metric]['AD'], cross_group_experiment[metric]['MCI'], cross_group_experiment[metric]['CN'])
            print("experiment", statistic, p_value)
            groups = np.array(['AD'] * len(cross_group_experiment[metric]['AD']) + ['MCI'] * len(cross_group_experiment[metric]['MCI']) + ['CN'] * len(cross_group_experiment[metric]['CN']))
            df = pd.DataFrame({'Data': np.concatenate([cross_group_experiment[metric]['AD'], cross_group_experiment[metric]['MCI'], cross_group_experiment[metric]['CN']]), 'Group': groups})
            model = ols('Data ~ C(Group)', data=df).fit()
            anova_table = sm.stats.anova_lm(model, typ=2)
            print(anova_table)
            if anova_table['PR(>F)']['C(Group)'] < 0.05:
                print("There is a significant difference in means among the groups.")
            else:
                print("There is no significant difference in means among the groups.")     
            posthoc_results = pairwise_tukeyhsd(endog=df['Data'], groups=df['Group'], alpha=0.05)
            print(posthoc_results) 
            
            
     